from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Optional
from datetime import datetime
import json
from pathlib import Path
import io
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Warning: openpyxl not installed. Excel features will not work.")

app = FastAPI(title="DCF Finanzanalyse API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://localhost:5173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

DATA_FILE = Path("data/analyses.json")
DATA_FILE.parent.mkdir(exist_ok=True)

class HistoricalData(BaseModel):
    revenue: List[float]
    ebit: List[float]
    depreciation: List[float]
    nwc_change: List[float]
    capex: List[float]
    interest: List[float]
    leverage_repayment: List[float]

class PeerCompany(BaseModel):
    name: str
    revenue: float
    ebitda: float
    ebit: float
    eps: float
    enterprise_value: float
    current_price: float
    book_value_per_share: float

class MultiplierInput(BaseModel):
    peer_companies: List[PeerCompany]
    target_revenue: float
    target_ebitda: float
    target_ebit: float
    target_eps: float
    target_book_value_per_share: float
    market_value_debt: float
    shares_outstanding: float
    current_share_price: Optional[float] = None

class DCFInput(BaseModel):
    company_name: str
    historical_data: HistoricalData
    forecast_years: int
    revenue_growth_rate: float
    risk_free_rate: float
    beta: float
    market_return: float
    cost_of_debt: float
    debt_capital: float
    tax_rate: float
    terminal_growth: float
    shares_outstanding: float
    net_debt: float
    current_share_price: Optional[float] = None
    target_revenue: Optional[float] = 0
    target_ebitda: Optional[float] = 0
    target_ebit: Optional[float] = 0
    target_eps: Optional[float] = 0
    target_book_value_per_share: Optional[float] = 0
    market_value_debt: Optional[float] = 0
    multiplier_data: Optional[dict] = None
    # Notizen und Dokumentation
    general_notes: Optional[str] = None
    assumptions: Optional[List[str]] = None
    sources: Optional[List[str]] = None
    risks: Optional[List[str]] = None

class DCFResult(BaseModel):
    id: str
    company_name: str
    inputs: DCFInput
    free_cash_flows: List[float]
    fcfe: List[float]
    noplat: List[float]
    tax_shield: List[float]
    forecast_revenues: List[float]
    terminal_value: float
    enterprise_value: float
    equity_value: float
    share_price: float
    equity_value_equity_approach: float
    share_price_equity_approach: float
    wacc: float
    cost_of_equity: float
    ebit_margin: float
    capex_percent: float
    nwc_percent: float
    multiplier_analysis: Optional[dict] = None
    sensitivity_analysis: Optional[dict] = None
    scenario_analysis: Optional[dict] = None
    created_at: str
    updated_at: str

def calculate_capm(risk_free_rate: float, beta: float, market_return: float) -> float:
    market_risk_premium = market_return - risk_free_rate
    return risk_free_rate + beta * market_risk_premium

def calculate_wacc_iterative(cost_of_equity: float, cost_of_debt: float, 
                             debt_capital: float, tax_rate: float, 
                             fcf_forecast: List[float], terminal_value_initial: float,
                             max_iterations: int = 25) -> tuple:
    wacc = 0.08
    
    for iteration in range(max_iterations):
        pv_fcf = sum([fcf_forecast[i] / ((1 + wacc) ** (i + 1)) for i in range(len(fcf_forecast))])
        pv_terminal = terminal_value_initial / ((1 + wacc) ** len(fcf_forecast))
        enterprise_value = pv_fcf + pv_terminal
        equity_capital = enterprise_value - debt_capital
        
        if equity_capital <= 0:
            equity_capital = 1
        
        total_capital = equity_capital + debt_capital
        equity_weight = equity_capital / total_capital
        debt_weight = debt_capital / total_capital
        
        new_wacc = (equity_weight * cost_of_equity) + (debt_weight * cost_of_debt * (1 - tax_rate))
        
        if abs(new_wacc - wacc) < 0.0000001:
            break
        
        wacc = new_wacc
    
    return wacc, enterprise_value, equity_capital

def calculate_multiples(multiplier_input: MultiplierInput) -> dict:
    peers = multiplier_input.peer_companies
    
    # Berechne Multiples für jede Peer Company
    ev_sales_multiples = []
    ev_ebit_multiples = []
    ev_ebitda_multiples = []
    per_multiples = []
    pbr_multiples = []
    
    for peer in peers:
        if peer.revenue > 0:
            ev_sales_multiples.append(peer.enterprise_value / peer.revenue)
        if peer.ebit > 0:
            ev_ebit_multiples.append(peer.enterprise_value / peer.ebit)
        if peer.ebitda > 0:
            ev_ebitda_multiples.append(peer.enterprise_value / peer.ebitda)
        if peer.eps > 0:
            per_multiples.append(peer.current_price / peer.eps)
        if peer.book_value_per_share > 0:
            pbr_multiples.append(peer.current_price / peer.book_value_per_share)
    
    # Berechne Statistiken (Mean, Median, Min, Max)
    def calc_stats(values):
        if not values:
            return {"mean": 0, "median": 0, "min": 0, "max": 0}
        sorted_vals = sorted(values)
        n = len(sorted_vals)
        return {
            "mean": sum(values) / n,
            "median": sorted_vals[n//2] if n % 2 == 1 else (sorted_vals[n//2-1] + sorted_vals[n//2]) / 2,
            "min": min(values),
            "max": max(values)
        }
    
    ev_sales_stats = calc_stats(ev_sales_multiples)
    ev_ebit_stats = calc_stats(ev_ebit_multiples)
    ev_ebitda_stats = calc_stats(ev_ebitda_multiples)
    per_stats = calc_stats(per_multiples)
    pbr_stats = calc_stats(pbr_multiples)
    
    # Berechne Target Valuations
    # Entity Approach (EV Multiples)
    ev_sales_valuation = multiplier_input.target_revenue * ev_sales_stats["median"]
    ev_ebit_valuation = multiplier_input.target_ebit * ev_ebit_stats["median"]
    ev_ebitda_valuation = multiplier_input.target_ebitda * ev_ebitda_stats["median"]
    
    # Equity Approach (Equity Multiples)
    per_share_price = multiplier_input.target_eps * per_stats["median"]
    pbr_share_price = multiplier_input.target_book_value_per_share * pbr_stats["median"]
    
    # Berechne Equity Values aus EV Multiples
    ev_sales_equity = ev_sales_valuation - multiplier_input.market_value_debt
    ev_ebit_equity = ev_ebit_valuation - multiplier_input.market_value_debt
    ev_ebitda_equity = ev_ebitda_valuation - multiplier_input.market_value_debt
    
    # Berechne Share Prices
    ev_sales_share_price = ev_sales_equity / multiplier_input.shares_outstanding
    ev_ebit_share_price = ev_ebit_equity / multiplier_input.shares_outstanding
    ev_ebitda_share_price = ev_ebitda_equity / multiplier_input.shares_outstanding
    
    # Berechne tatsächliche Target Multiples (basierend auf aktuellen Marktdaten)
    # Nur wenn current_share_price vorhanden ist
    if multiplier_input.current_share_price and multiplier_input.current_share_price > 0:
        target_market_cap = multiplier_input.shares_outstanding * multiplier_input.current_share_price
        target_enterprise_value = target_market_cap + multiplier_input.market_value_debt
        
        target_ev_sales_actual = target_enterprise_value / multiplier_input.target_revenue if multiplier_input.target_revenue > 0 else 0
        target_ev_ebit_actual = target_enterprise_value / multiplier_input.target_ebit if multiplier_input.target_ebit > 0 else 0
        target_ev_ebitda_actual = target_enterprise_value / multiplier_input.target_ebitda if multiplier_input.target_ebitda > 0 else 0
        target_per_actual = multiplier_input.current_share_price / multiplier_input.target_eps if multiplier_input.target_eps > 0 else 0
        target_pbr_actual = multiplier_input.current_share_price / multiplier_input.target_book_value_per_share if multiplier_input.target_book_value_per_share > 0 else 0
    else:
        target_ev_sales_actual = None
        target_ev_ebit_actual = None
        target_ev_ebitda_actual = None
        target_per_actual = None
        target_pbr_actual = None
    
    # Durchschnitte
    mean_share_price = (ev_sales_share_price + ev_ebit_share_price + ev_ebitda_share_price + per_share_price + pbr_share_price) / 5
    median_share_price = sorted([ev_sales_share_price, ev_ebit_share_price, ev_ebitda_share_price, per_share_price, pbr_share_price])[2]
    
    return {
        "multiples_stats": {
            "ev_sales": ev_sales_stats,
            "ev_ebit": ev_ebit_stats,
            "ev_ebitda": ev_ebitda_stats,
            "per": per_stats,
            "pbr": pbr_stats
        },
        "target_multiples": {
            "ev_sales": round(target_ev_sales_actual, 2) if target_ev_sales_actual is not None else None,
            "ev_ebit": round(target_ev_ebit_actual, 2) if target_ev_ebit_actual is not None else None,
            "ev_ebitda": round(target_ev_ebitda_actual, 2) if target_ev_ebitda_actual is not None else None,
            "per": round(target_per_actual, 2) if target_per_actual is not None else None,
            "pbr": round(target_pbr_actual, 2) if target_pbr_actual is not None else None
        } if target_ev_sales_actual is not None else None,
        "valuations": {
            "ev_sales": round(ev_sales_valuation, 2),
            "ev_ebit": round(ev_ebit_valuation, 2),
            "ev_ebitda": round(ev_ebitda_valuation, 2),
            "ev_sales_equity": round(ev_sales_equity, 2),
            "ev_ebit_equity": round(ev_ebit_equity, 2),
            "ev_ebitda_equity": round(ev_ebitda_equity, 2)
        },
        "share_prices": {
            "ev_sales": round(ev_sales_share_price, 2),
            "ev_ebit": round(ev_ebit_share_price, 2),
            "ev_ebitda": round(ev_ebitda_share_price, 2),
            "per": round(per_share_price, 2),
            "pbr": round(pbr_share_price, 2),
            "mean": round(mean_share_price, 2),
            "median": round(median_share_price, 2)
        }
    }

def calculate_dcf(inputs: DCFInput, skip_analysis: bool = False) -> dict:
    hist = inputs.historical_data
    
    historical_revenue_avg = sum(hist.revenue) / len(hist.revenue)
    ebit_margin_avg = sum([hist.ebit[i] / hist.revenue[i] for i in range(len(hist.revenue))]) / len(hist.revenue)
    da_percent_avg = sum([hist.depreciation[i] / hist.revenue[i] for i in range(len(hist.revenue))]) / len(hist.revenue)
    nwc_percent_avg = sum([hist.nwc_change[i] / hist.revenue[i] for i in range(len(hist.revenue))]) / len(hist.revenue)
    capex_percent_avg = sum([hist.capex[i] / hist.revenue[i] for i in range(len(hist.revenue))]) / len(hist.revenue)
    interest_percent_avg = sum([hist.interest[i] / hist.revenue[i] for i in range(len(hist.revenue))]) / len(hist.revenue)
    leverage_percent_avg = sum([hist.leverage_repayment[i] / hist.revenue[i] for i in range(len(hist.revenue))]) / len(hist.revenue)
    
    last_revenue = hist.revenue[-1]
    forecast_revenues = []
    for i in range(inputs.forecast_years):
        forecast_revenues.append(last_revenue * ((1 + inputs.revenue_growth_rate) ** (i + 1)))
    
    fcf_forecast = []
    fcfe_forecast = []
    tax_shield_forecast = []
    noplat_forecast = []
    
    for i, revenue in enumerate(forecast_revenues):
        ebit = revenue * ebit_margin_avg
        tax = ebit * inputs.tax_rate
        noplat = ebit - tax
        noplat_forecast.append(noplat)
        
        depreciation = revenue * da_percent_avg
        nwc_change = revenue * nwc_percent_avg
        capex = revenue * capex_percent_avg
        
        fcf = noplat + depreciation - nwc_change - capex
        fcf_forecast.append(fcf)
        
        interest = revenue * interest_percent_avg
        tax_shield = interest * inputs.tax_rate
        tax_shield_forecast.append(tax_shield)
        
        leverage_repayment = revenue * leverage_percent_avg
        fcfe = fcf + tax_shield - interest - leverage_repayment
        fcfe_forecast.append(fcfe)
    
    cost_of_equity = calculate_capm(inputs.risk_free_rate, inputs.beta, inputs.market_return)
    
    terminal_fcf = fcf_forecast[-1] * (1 + inputs.terminal_growth)
    terminal_value_initial = terminal_fcf / (0.08 - inputs.terminal_growth)
    
    wacc, enterprise_value, equity_capital = calculate_wacc_iterative(
        cost_of_equity, inputs.cost_of_debt, inputs.debt_capital, 
        inputs.tax_rate, fcf_forecast, terminal_value_initial
    )
    
    terminal_value = terminal_fcf / (wacc - inputs.terminal_growth)
    pv_fcf = sum([fcf_forecast[i] / ((1 + wacc) ** (i + 1)) for i in range(len(fcf_forecast))])
    pv_terminal = terminal_value / ((1 + wacc) ** len(fcf_forecast))
    enterprise_value = pv_fcf + pv_terminal
    
    equity_value = enterprise_value - inputs.net_debt
    share_price = equity_value / inputs.shares_outstanding
    
    terminal_fcfe = fcfe_forecast[-1] * (1 + inputs.terminal_growth)
    terminal_value_equity = terminal_fcfe / (cost_of_equity - inputs.terminal_growth)
    pv_fcfe = sum([fcfe_forecast[i] / ((1 + cost_of_equity) ** (i + 1)) for i in range(len(fcfe_forecast))])
    pv_terminal_equity = terminal_value_equity / ((1 + cost_of_equity) ** len(fcfe_forecast))
    equity_value_equity_approach = pv_fcfe + pv_terminal_equity
    share_price_equity_approach = equity_value_equity_approach / inputs.shares_outstanding
    
    result = {
        "free_cash_flows": [round(f, 2) for f in fcf_forecast],
        "fcfe": [round(f, 2) for f in fcfe_forecast],
        "noplat": [round(n, 2) for n in noplat_forecast],
        "tax_shield": [round(t, 2) for t in tax_shield_forecast],
        "forecast_revenues": [round(r, 2) for r in forecast_revenues],
        "terminal_value": round(terminal_value, 2),
        "enterprise_value": round(enterprise_value, 2),
        "equity_value": round(equity_value, 2),
        "share_price": round(share_price, 2),
        "equity_value_equity_approach": round(equity_value_equity_approach, 2),
        "share_price_equity_approach": round(share_price_equity_approach, 2),
        "wacc": round(wacc, 4),
        "cost_of_equity": round(cost_of_equity, 4),
        "ebit_margin": round(ebit_margin_avg, 4),
        "capex_percent": round(capex_percent_avg, 4),
        "nwc_percent": round(nwc_percent_avg, 4)
    }
    
    # Füge Multiplier-Analyse hinzu, falls vorhanden
    if inputs.multiplier_data and inputs.multiplier_data.get('peer_companies'):
        # Konstruiere MultiplierInput mit Daten von oberster Ebene
        peer_companies = [PeerCompany(**peer) for peer in inputs.multiplier_data['peer_companies']]
        multiplier_input = MultiplierInput(
            peer_companies=peer_companies,
            target_revenue=inputs.target_revenue or 0,
            target_ebitda=inputs.target_ebitda or 0,
            target_ebit=inputs.target_ebit or 0,
            target_eps=inputs.target_eps or 0,
            target_book_value_per_share=inputs.target_book_value_per_share or 0,
            market_value_debt=inputs.market_value_debt or 0,
            shares_outstanding=inputs.shares_outstanding,
            current_share_price=inputs.current_share_price
        )
        multiplier_results = calculate_multiples(multiplier_input)
        result["multiplier_analysis"] = multiplier_results
    
    # Berechne Sensitivitäts- und Szenarioanalyse nur wenn nicht übersprungen
    if not skip_analysis:
        try:
            sensitivity_data = calculate_sensitivity_internal(inputs, result)
            result["sensitivity_analysis"] = sensitivity_data
        except Exception as e:
            print(f"Sensitivity analysis error: {e}")
            result["sensitivity_analysis"] = None
        
        try:
            scenario_data = calculate_scenarios_internal(inputs, result)
            result["scenario_analysis"] = scenario_data
        except Exception as e:
            print(f"Scenario analysis error: {e}")
            result["scenario_analysis"] = None
    
    return result

def load_analyses():
    if DATA_FILE.exists():
        with open(DATA_FILE, 'r') as f:
            return json.load(f)
    return []

def save_analyses(analyses):
    with open(DATA_FILE, 'w') as f:
        json.dump(analyses, f, indent=2)

@app.get("/")
def root():
    return {"message": "DCF Finanzanalyse API"}

@app.post("/api/analyses", response_model=DCFResult)
def create_analysis(inputs: DCFInput):
    results = calculate_dcf(inputs)
    
    analyses = load_analyses()
    analysis_id = f"dcf_{len(analyses) + 1}_{datetime.now().strftime('%Y%m%d%H%M%S')}"
    
    analysis = {
        "id": analysis_id,
        "company_name": inputs.company_name,
        "inputs": inputs.model_dump(),
        **results,
        "created_at": datetime.now().isoformat(),
        "updated_at": datetime.now().isoformat()
    }
    
    analyses.append(analysis)
    save_analyses(analyses)
    
    return analysis

@app.get("/api/analyses", response_model=List[DCFResult])
def get_analyses():
    analyses = load_analyses()
    # Migration: Konvertiere alte multiplier_analysis Struktur
    for analysis in analyses:
        if "multiplier_analysis" in analysis:
            ma = analysis["multiplier_analysis"]
            # Alte Struktur hatte target_multiples_implied und target_multiples_actual
            if "target_multiples_implied" in ma or "target_multiples_actual" in ma:
                # Verwende target_multiples_actual wenn vorhanden, sonst None
                if "target_multiples_actual" in ma:
                    ma["target_multiples"] = ma["target_multiples_actual"]
                else:
                    ma["target_multiples"] = None
                # Entferne alte Keys
                ma.pop("target_multiples_implied", None)
                ma.pop("target_multiples_actual", None)
    return analyses

@app.get("/api/analyses/{analysis_id}", response_model=DCFResult)
def get_analysis(analysis_id: str):
    analyses = load_analyses()
    for analysis in analyses:
        if analysis["id"] == analysis_id:
            # Migration: Konvertiere alte multiplier_analysis Struktur
            if "multiplier_analysis" in analysis:
                ma = analysis["multiplier_analysis"]
                # Alte Struktur hatte target_multiples_implied und target_multiples_actual
                if "target_multiples_implied" in ma or "target_multiples_actual" in ma:
                    # Verwende target_multiples_actual wenn vorhanden, sonst None
                    if "target_multiples_actual" in ma:
                        ma["target_multiples"] = ma["target_multiples_actual"]
                    else:
                        ma["target_multiples"] = None
                    # Entferne alte Keys
                    ma.pop("target_multiples_implied", None)
                    ma.pop("target_multiples_actual", None)
            return analysis
    raise HTTPException(status_code=404, detail="Analysis not found")

@app.put("/api/analyses/{analysis_id}", response_model=DCFResult)
def update_analysis(analysis_id: str, inputs: DCFInput):
    analyses = load_analyses()
    
    for i, analysis in enumerate(analyses):
        if analysis["id"] == analysis_id:
            results = calculate_dcf(inputs)
            
            analyses[i] = {
                "id": analysis_id,
                "company_name": inputs.company_name,
                "inputs": inputs.model_dump(),
                **results,
                "created_at": analysis["created_at"],
                "updated_at": datetime.now().isoformat()
            }
            
            save_analyses(analyses)
            return analyses[i]
    
    raise HTTPException(status_code=404, detail="Analysis not found")

@app.delete("/api/analyses/{analysis_id}")
def delete_analysis(analysis_id: str):
    analyses = load_analyses()
    analyses = [a for a in analyses if a["id"] != analysis_id]
    save_analyses(analyses)
    return {"message": "Analysis deleted"}

@app.get("/api/excel-template")
def download_excel_template():
    """Generiere und sende Excel-Vorlage für DCF-Eingaben"""
    wb = Workbook()
    ws = wb.active
    ws.title = "DCF Eingaben"
    
    # Header-Styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    input_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Gelb für Eingabefelder
    
    # Unternehmensparameter
    ws['A1'] = 'UNTERNEHMENSPARAMETER'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    ws['A1'].font = Font(bold=True, color="FFFFFF", size=14)
    
    ws['A2'] = 'Unternehmensname'
    ws['B2'] = 'Beispiel AG'
    ws['B2'].fill = input_fill
    ws['A3'] = 'Historisches Startjahr'
    ws['B3'] = 2020
    ws['B3'].fill = input_fill
    ws['A4'] = 'Aktien ausstehend (Mio.)'
    ws['B4'] = 100
    ws['B4'].fill = input_fill
    ws['A5'] = 'Net Debt (Mio. €)'
    ws['B5'] = 0
    ws['B5'].fill = input_fill
    ws['A6'] = 'Aktueller Marktpreis (€)'
    ws['B6'] = 100
    ws['B6'].fill = input_fill
    
    # Aktuelle Finanzkennzahlen (für Multiplier-Bewertung)
    ws['A8'] = 'AKTUELLE FINANZKENNZAHLEN (für Multiplier-Bewertung)'
    ws['A8'].font = Font(bold=True, size=14)
    ws['A8'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    ws['A8'].font = Font(bold=True, color="FFFFFF", size=14)
    
    current_financials = [
        ('Umsatz (Mio. €)', 0),
        ('EBITDA (Mio. €)', 0),
        ('EBIT (Mio. €)', 0),
        ('EPS (€)', 0),
        ('Buchwert/Aktie (€)', 0),
        ('Marktwert Schulden (Mio. €)', 0)
    ]
    
    row = 9
    for param, value in current_financials:
        ws[f'A{row}'] = param
        ws[f'B{row}'] = value
        ws[f'B{row}'].fill = input_fill
        row += 1
    
    # Historische Daten
    ws['A16'] = 'HISTORISCHE DATEN (Mio. €)'
    ws['A16'].font = Font(bold=True, size=14)
    ws['A16'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    ws['A16'].font = Font(bold=True, color="FFFFFF", size=14)
    
    # Erklärung Zeitbezug
    ws.merge_cells('A15:K15')
    ws['A15'] = 'Hinweis: Startjahr in B3 ändern → Jahre aktualisieren sich automatisch. Mindestens 4 Jahre erforderlich. Neuestes Jahr = t, davor = t-1, t-2, t-3, etc.'
    ws['A15'].font = Font(italic=True, size=9, color="0066CC")
    ws['A15'].alignment = Alignment(horizontal='center')
    
    # Header-Zeilen für Jahre
    ws['A17'] = 'Jahr'
    ws['A18'] = 'Zeitbezug'
    ws['A19'] = 'Kennzahl'
    
    year_cols = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
    
    # Zeile 17: Tatsächliche Jahre - MIT FORMELN für automatische Aktualisierung
    for i, col in enumerate(year_cols):
        # Formel: Startjahr (B3) + Offset
        ws[f'{col}17'] = f'=$B$3+{i}'
        ws[f'{col}17'].fill = header_fill
        ws[f'{col}17'].font = header_font
        ws[f'{col}17'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Zeile 18: Zeitbezüge (Beispiel für 4 Jahre)
    # Wenn 4 Jahre: B=t-3, C=t-2, D=t-1, E=t, F-K leer
    time_refs_example = ['t-3', 't-2', 't-1', 't', '', '', '', '', '', '']
    for i, col in enumerate(year_cols):
        ws[f'{col}18'] = time_refs_example[i]
        ws[f'{col}18'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        ws[f'{col}18'].font = Font(italic=True, size=9, color="666666")
        ws[f'{col}18'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Formatiere Header-Zeilen
    for cell in ['A17', 'A18', 'A19']:
        ws[cell].fill = header_fill
        ws[cell].font = header_font
        ws[cell].alignment = Alignment(horizontal='center', vertical='center')
    
    # Beispieldaten für 4 Jahre (Rest bleibt leer für flexible Eingabe)
    historical_fields = [
        ('Umsatz', [140500, 148500, 152671, 160000]),
        ('EBIT', [10751, 13400, 14164, 15200]),
        ('Abschreibungen', [260, 307, 353, 380]),
        ('NWC Veränderung', [-2318, -1385, -187, -150]),
        ('CapEx', [1501, 1400, 1444, 1500]),
        ('Zinsen', [999, 1160, 1438, 1500]),
        ('Tilgung/Aufnahme FK', [-2967, -201, 1057, 800])
    ]
    
    row = 20  # Jetzt ab Zeile 20 (wegen verschobener Header)
    for field, values in historical_fields:
        ws[f'A{row}'] = field
        # Fülle Beispieldaten für erste 3 Jahre
        for i, val in enumerate(values):
            ws[f'{year_cols[i]}{row}'] = val
            ws[f'{year_cols[i]}{row}'].fill = input_fill
        # Markiere restliche Zellen auch als Eingabefelder
        for i in range(len(values), 10):
            ws[f'{year_cols[i]}{row}'].fill = input_fill
        row += 1
    
    # CAPM & Bewertungsparameter (jetzt ab Zeile 28)
    ws['A28'] = 'CAPM & BEWERTUNGSPARAMETER'
    ws['A28'].font = Font(bold=True, size=14)
    ws['A28'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    ws['A28'].font = Font(bold=True, color="FFFFFF", size=14)
    
    params = [
        ('Risikofreier Zinssatz', 0.025),
        ('Beta', 1.0),
        ('Marktrendite', 0.08),
        ('Fremdkapitalkosten', 0.03),
        ('Fremdkapital (Mio. €)', 0),
        ('Steuersatz', 0.30),
        ('Umsatzwachstum (Prognose)', 0.03),
        ('Terminal Growth', 0.02),
        ('Prognosejahre', 4)
    ]
    
    row = 29  # Angepasst
    for param, value in params:
        ws[f'A{row}'] = param
        ws[f'B{row}'] = value
        ws[f'B{row}'].fill = input_fill
        row += 1
    
    # MULTIPLIER-DATEN (Optional)
    ws['A40'] = 'MULTIPLIER-DATEN (Optional)'
    ws['A40'].font = Font(bold=True, size=14)
    ws['A40'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    ws['A40'].font = Font(bold=True, color="FFFFFF", size=14)
    
    multiplier_params = [
        ('EV/Sales Multiple', None),
        ('EV/EBIT Multiple', None),
        ('EV/EBITDA Multiple', None),
        ('P/E Ratio', None),
        ('Price/Book Ratio', None)
    ]
    
    row = 41  # Angepasst
    for param, value in multiplier_params:
        ws[f'A{row}'] = param
        ws[f'B{row}'] = value if value else ''
        ws[f'B{row}'].fill = input_fill
        row += 1
    
    # Spaltenbreiten anpassen
    ws.column_dimensions['A'].width = 30
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
        ws.column_dimensions[col].width = 15
    
    # Hinweise
    ws['A48'] = 'HINWEISE:'
    ws['A48'].font = Font(bold=True, size=12, color="C00000")
    ws['A49'] = '• Füllen Sie alle gelben Felder aus'
    ws['A50'] = '• Aktuelle Finanzkennzahlen: Für Multiplier-Bewertung des Target-Unternehmens'
    ws['A51'] = '• Historische Daten: 4-10 Jahre (mindestens 4 erforderlich)'
    ws['A52'] = '• Zeitbezug: Letztes ausgefülltes Jahr = t (aktuell), davor = t-1, t-2, etc.'
    ws['A53'] = '• Jahre ändern sich automatisch wenn Sie Startjahr (B3) anpassen'
    ws['A54'] = '• Alle Geldbeträge in Mio. €'
    ws['A55'] = '• Prozentangaben als Dezimalzahl (z.B. 0.03 für 3%)'
    ws['A56'] = '• Multiplier-Daten sind optional'
    ws['A57'] = '• Leere Jahresspalten werden ignoriert'
    
    # Speichere in BytesIO
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=DCF_Vorlage.xlsx"}
    )

@app.post("/api/upload-excel")
async def upload_excel(file: UploadFile = File(...)):
    """Parse Excel-Datei und erstelle DCF-Eingaben mit Validierung"""
    errors = []
    warnings = []
    
    try:
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents))
        ws = wb.active
        
        # Parse Unternehmensparameter mit Validierung
        company_name = ws['B2'].value
        if not company_name or str(company_name).strip() == '':
            errors.append("Unternehmensname ist erforderlich (Zelle B2)")
            company_name = 'Unbekannt'
        
        try:
            historical_start_year = int(ws['B3'].value or 2020)
            if historical_start_year < 1900 or historical_start_year > 2100:
                warnings.append(f"Ungewöhnliches Startjahr: {historical_start_year}")
        except (ValueError, TypeError):
            errors.append("Historisches Startjahr muss eine Zahl sein (Zelle B3)")
            historical_start_year = 2020
        
        try:
            shares_outstanding = float(ws['B4'].value or 100)
            if shares_outstanding <= 0:
                errors.append("Aktien ausstehend muss größer als 0 sein (Zelle B4)")
        except (ValueError, TypeError):
            errors.append("Aktien ausstehend muss eine Zahl sein (Zelle B4)")
            shares_outstanding = 100
        
        try:
            net_debt = float(ws['B5'].value or 0)
        except (ValueError, TypeError):
            warnings.append("Net Debt konnte nicht gelesen werden, verwende 0 (Zelle B5)")
            net_debt = 0
        
        current_share_price = ws['B6'].value
        if current_share_price:
            try:
                current_share_price = float(current_share_price)
                if current_share_price <= 0:
                    warnings.append("Aktienkurs sollte positiv sein (Zelle B6)")
            except (ValueError, TypeError):
                warnings.append("Aktienkurs konnte nicht gelesen werden (Zelle B6)")
                current_share_price = None
        
        # Parse aktuelle Finanzkennzahlen (Zeilen 9-14)
        try:
            target_revenue = float(ws['B9'].value or 0)
        except (ValueError, TypeError):
            warnings.append("Aktueller Umsatz konnte nicht gelesen werden (Zelle B9)")
            target_revenue = 0
        
        try:
            target_ebitda = float(ws['B10'].value or 0)
        except (ValueError, TypeError):
            warnings.append("Aktuelles EBITDA konnte nicht gelesen werden (Zelle B10)")
            target_ebitda = 0
        
        try:
            target_ebit = float(ws['B11'].value or 0)
        except (ValueError, TypeError):
            warnings.append("Aktuelles EBIT konnte nicht gelesen werden (Zelle B11)")
            target_ebit = 0
        
        try:
            target_eps = float(ws['B12'].value or 0)
        except (ValueError, TypeError):
            warnings.append("Aktuelles EPS konnte nicht gelesen werden (Zelle B12)")
            target_eps = 0
        
        try:
            target_book_value_per_share = float(ws['B13'].value or 0)
        except (ValueError, TypeError):
            warnings.append("Aktueller Buchwert/Aktie konnte nicht gelesen werden (Zelle B13)")
            target_book_value_per_share = 0
        
        try:
            market_value_debt = float(ws['B14'].value or 0)
        except (ValueError, TypeError):
            warnings.append("Marktwert Schulden konnte nicht gelesen werden (Zelle B14)")
            market_value_debt = 0
        
        # Parse historische Daten - flexible Anzahl Jahre (3-10)
        year_cols = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
        field_rows = {
            'revenue': 20,  # Angepasst auf neue Struktur
            'ebit': 21,
            'depreciation': 22,
            'nwc_change': 23,
            'capex': 24,
            'interest': 25,
            'leverage_repayment': 26
        }
        
        historical_data = {field: [] for field in field_rows.keys()}
        
        # Bestimme Anzahl der Jahre durch Prüfung der Umsatz-Zeile (jetzt Zeile 20)
        num_years = 0
        for col in year_cols:
            val = ws[f'{col}20'].value  # Angepasst auf neue Struktur
            if val is not None and val != '':
                num_years += 1
            else:
                break
        
        if num_years < 4:
            errors.append(f"Mindestens 4 historische Jahre erforderlich, gefunden: {num_years}")
            num_years = 4
        elif num_years > 10:
            warnings.append(f"Mehr als 10 Jahre gefunden, verwende nur die ersten 10")
            num_years = 10
        
        # Parse Daten für alle Felder
        for field, row_num in field_rows.items():
            for i in range(num_years):
                try:
                    val = ws[f'{year_cols[i]}{row_num}'].value
                    if val is None or val == '':
                        warnings.append(f"{field} Jahr {i+1} ist leer (Zelle {year_cols[i]}{row_num}), verwende 0")
                        val = 0
                    historical_data[field].append(float(val))
                except (ValueError, TypeError):
                    errors.append(f"Ungültiger Wert für {field} Jahr {i+1} (Zelle {year_cols[i]}{row_num})")
                    historical_data[field].append(0)
        
        # Parse CAPM & Bewertungsparameter mit Validierung (jetzt ab Zeile 29)
        try:
            risk_free_rate = float(ws['B29'].value or 0.025)
            if risk_free_rate < 0 or risk_free_rate > 0.2:
                warnings.append(f"Ungewöhnlicher risikofreier Zinssatz: {risk_free_rate}")
        except (ValueError, TypeError):
            errors.append("Risikofreier Zinssatz muss eine Zahl sein (Zelle B29)")
            risk_free_rate = 0.025
        
        try:
            beta = float(ws['B30'].value or 1.0)
            if beta < 0:
                warnings.append("Beta sollte nicht negativ sein")
        except (ValueError, TypeError):
            errors.append("Beta muss eine Zahl sein (Zelle B30)")
            beta = 1.0
        
        try:
            market_return = float(ws['B31'].value or 0.08)
        except (ValueError, TypeError):
            errors.append("Marktrendite muss eine Zahl sein (Zelle B31)")
            market_return = 0.08
        
        try:
            cost_of_debt = float(ws['B32'].value or 0.03)
        except (ValueError, TypeError):
            warnings.append("Fremdkapitalkosten konnte nicht gelesen werden (Zelle B32)")
            cost_of_debt = 0.03
        
        try:
            debt_capital = float(ws['B33'].value or 0)
        except (ValueError, TypeError):
            warnings.append("Fremdkapital konnte nicht gelesen werden (Zelle B33)")
            debt_capital = 0
        
        try:
            tax_rate = float(ws['B34'].value or 0.30)
            if tax_rate < 0 or tax_rate > 1:
                errors.append("Steuersatz muss zwischen 0 und 1 liegen (Zelle B34)")
        except (ValueError, TypeError):
            errors.append("Steuersatz muss eine Zahl sein (Zelle B34)")
            tax_rate = 0.30
        
        try:
            revenue_growth_rate = float(ws['B35'].value or 0.03)
        except (ValueError, TypeError):
            warnings.append("Umsatzwachstum konnte nicht gelesen werden (Zelle B35)")
            revenue_growth_rate = 0.03
        
        try:
            terminal_growth = float(ws['B36'].value or 0.02)
            if terminal_growth < 0 or terminal_growth > 0.1:
                warnings.append(f"Ungewöhnliches Terminal Growth: {terminal_growth}")
        except (ValueError, TypeError):
            errors.append("Terminal Growth muss eine Zahl sein (Zelle B36)")
            terminal_growth = 0.02
        
        try:
            forecast_years = int(ws['B37'].value or 4)
            if forecast_years < 1 or forecast_years > 20:
                warnings.append(f"Ungewöhnliche Anzahl Prognosejahre: {forecast_years}")
        except (ValueError, TypeError):
            errors.append("Prognosejahre muss eine ganze Zahl sein (Zelle B37)")
            forecast_years = 4
        
        # Parse Multiplier-Daten (Optional) - jetzt ab Zeile 41
        multiplier_data = None
        try:
            ev_sales = ws['B41'].value
            ev_ebit = ws['B42'].value
            ev_ebitda = ws['B43'].value
            pe_ratio = ws['B44'].value
            pb_ratio = ws['B45'].value
            
            if any([ev_sales, ev_ebit, ev_ebitda, pe_ratio, pb_ratio]):
                multiplier_data = {}
                if ev_sales: multiplier_data['ev_sales'] = float(ev_sales)
                if ev_ebit: multiplier_data['ev_ebit'] = float(ev_ebit)
                if ev_ebitda: multiplier_data['ev_ebitda'] = float(ev_ebitda)
                if pe_ratio: multiplier_data['per'] = float(pe_ratio)
                if pb_ratio: multiplier_data['pbr'] = float(pb_ratio)
        except (ValueError, TypeError):
            warnings.append("Einige Multiplier-Daten konnten nicht gelesen werden")
        
        # Wenn kritische Fehler vorhanden sind, Fehler werfen
        if errors:
            error_msg = "Validierungsfehler:\n" + "\n".join(f"• {e}" for e in errors)
            if warnings:
                error_msg += "\n\nWarnungen:\n" + "\n".join(f"• {w}" for w in warnings)
            raise HTTPException(status_code=400, detail=error_msg)
        
        # Erstelle DCFInput
        dcf_input = {
            'company_name': company_name,
            'historical_start_year': historical_start_year,
            'historical_data': historical_data,
            'forecast_years': forecast_years,
            'revenue_growth_rate': revenue_growth_rate,
            'risk_free_rate': risk_free_rate,
            'beta': beta,
            'market_return': market_return,
            'cost_of_debt': cost_of_debt,
            'debt_capital': debt_capital,
            'tax_rate': tax_rate,
            'terminal_growth': terminal_growth,
            'shares_outstanding': shares_outstanding,
            'net_debt': net_debt,
            'current_share_price': current_share_price,
            'target_revenue': target_revenue,
            'target_ebitda': target_ebitda,
            'target_ebit': target_ebit,
            'target_eps': target_eps,
            'target_book_value_per_share': target_book_value_per_share,
            'market_value_debt': market_value_debt,
            'multiplier_data': multiplier_data
        }
        
        # Füge Warnungen als Metadaten hinzu
        response_data = {
            'data': dcf_input,
            'warnings': warnings if warnings else None,
            'info': f"✓ {num_years} historische Jahre geladen"
        }
        
        return response_data
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Fehler beim Parsen der Excel-Datei: {str(e)}")

def calculate_sensitivity_internal(inputs: DCFInput, base_results: dict) -> dict:
    """Interne Funktion für Sensitivitätsanalyse - wird automatisch bei DCF-Berechnung aufgerufen"""
    # WACC Range: ±2%
    wacc_range = [-0.02, -0.015, -0.01, -0.005, 0, 0.005, 0.01, 0.015, 0.02]
    # Terminal Growth Range: ±1%
    growth_range = [-0.01, -0.0075, -0.005, -0.0025, 0, 0.0025, 0.005, 0.0075, 0.01]
    
    sensitivity_matrix = []
    
    for wacc_delta in wacc_range:
        row = []
        for growth_delta in growth_range:
            modified_inputs = inputs.model_copy(deep=True)
            modified_inputs.risk_free_rate = inputs.risk_free_rate + wacc_delta
            modified_inputs.terminal_growth = inputs.terminal_growth + growth_delta
            
            try:
                result = calculate_dcf(modified_inputs, skip_analysis=True)
                row.append({
                    "wacc_delta": wacc_delta,
                    "growth_delta": growth_delta,
                    "share_price": result["share_price"],
                    "wacc": result["wacc"],
                    "terminal_growth": modified_inputs.terminal_growth
                })
            except:
                row.append({
                    "wacc_delta": wacc_delta,
                    "growth_delta": growth_delta,
                    "share_price": 0,
                    "wacc": 0,
                    "terminal_growth": 0
                })
        
        sensitivity_matrix.append(row)
    
    return {
        "base_share_price": base_results["share_price"],
        "base_wacc": base_results["wacc"],
        "base_terminal_growth": inputs.terminal_growth,
        "sensitivity_matrix": sensitivity_matrix,
        "wacc_range": wacc_range,
        "growth_range": growth_range
    }

def calculate_scenarios_internal(inputs: DCFInput, base_results: dict) -> dict:
    """Interne Funktion für Szenarioanalyse - wird automatisch bei DCF-Berechnung aufgerufen"""
    # Best Case: +20% Wachstum, -0.5% WACC
    best_inputs = inputs.model_copy(deep=True)
    best_inputs.revenue_growth_rate = inputs.revenue_growth_rate * 1.2
    best_inputs.terminal_growth = min(inputs.terminal_growth * 1.2, 0.05)
    best_inputs.risk_free_rate = max(inputs.risk_free_rate - 0.005, 0.01)
    best_results = calculate_dcf(best_inputs, skip_analysis=True)
    
    # Worst Case: -20% Wachstum, +0.5% WACC
    worst_inputs = inputs.model_copy(deep=True)
    worst_inputs.revenue_growth_rate = inputs.revenue_growth_rate * 0.8
    worst_inputs.terminal_growth = max(inputs.terminal_growth * 0.8, 0.005)
    worst_inputs.risk_free_rate = inputs.risk_free_rate + 0.005
    worst_results = calculate_dcf(worst_inputs, skip_analysis=True)
    
    return {
        "base_case": {
            "name": "Base Case",
            "share_price": base_results["share_price"],
            "enterprise_value": base_results["enterprise_value"],
            "wacc": base_results["wacc"],
            "revenue_growth": inputs.revenue_growth_rate,
            "terminal_growth": inputs.terminal_growth
        },
        "best_case": {
            "name": "Best Case",
            "share_price": best_results["share_price"],
            "enterprise_value": best_results["enterprise_value"],
            "wacc": best_results["wacc"],
            "revenue_growth": best_inputs.revenue_growth_rate,
            "terminal_growth": best_inputs.terminal_growth
        },
        "worst_case": {
            "name": "Worst Case",
            "share_price": worst_results["share_price"],
            "enterprise_value": worst_results["enterprise_value"],
            "wacc": worst_results["wacc"],
            "revenue_growth": worst_inputs.revenue_growth_rate,
            "terminal_growth": worst_inputs.terminal_growth
        }
    }

@app.post("/api/sensitivity")
def calculate_sensitivity(inputs: DCFInput):
    """Berechne Sensitivitätsanalyse für WACC und Terminal Growth"""
    base_results = calculate_dcf(inputs)
    
    # WACC Range: ±2%
    wacc_range = [-0.02, -0.015, -0.01, -0.005, 0, 0.005, 0.01, 0.015, 0.02]
    # Terminal Growth Range: ±1%
    growth_range = [-0.01, -0.0075, -0.005, -0.0025, 0, 0.0025, 0.005, 0.0075, 0.01]
    
    sensitivity_matrix = []
    
    for wacc_delta in wacc_range:
        row = []
        for growth_delta in growth_range:
            # Erstelle modifizierte Inputs
            modified_inputs = inputs.model_copy(deep=True)
            
            # Berechne mit modifizierten Parametern
            # WACC wird in calculate_dcf berechnet, daher müssen wir die Basis-Parameter anpassen
            # Für Vereinfachung: Passe risk_free_rate an (beeinflusst WACC)
            modified_inputs.risk_free_rate = inputs.risk_free_rate + wacc_delta
            modified_inputs.terminal_growth = inputs.terminal_growth + growth_delta
            
            try:
                result = calculate_dcf(modified_inputs, skip_analysis=True)
                row.append({
                    "wacc_delta": wacc_delta,
                    "growth_delta": growth_delta,
                    "share_price": result["share_price"],
                    "wacc": result["wacc"],
                    "terminal_growth": modified_inputs.terminal_growth
                })
            except:
                row.append({
                    "wacc_delta": wacc_delta,
                    "growth_delta": growth_delta,
                    "share_price": 0,
                    "wacc": 0,
                    "terminal_growth": 0
                })
        
        sensitivity_matrix.append(row)
    
    return {
        "base_share_price": base_results["share_price"],
        "base_wacc": base_results["wacc"],
        "base_terminal_growth": inputs.terminal_growth,
        "sensitivity_matrix": sensitivity_matrix,
        "wacc_range": wacc_range,
        "growth_range": growth_range
    }

@app.post("/api/scenarios")
def calculate_scenarios(inputs: DCFInput):
    """Berechne Best/Base/Worst Case Szenarien"""
    
    # Base Case
    base_results = calculate_dcf(inputs)
    
    # Best Case: +20% Wachstum, -0.5% WACC
    best_inputs = inputs.model_copy(deep=True)
    best_inputs.revenue_growth_rate = inputs.revenue_growth_rate * 1.2
    best_inputs.terminal_growth = min(inputs.terminal_growth * 1.2, 0.05)  # Max 5%
    best_inputs.risk_free_rate = max(inputs.risk_free_rate - 0.005, 0.01)  # Min 1%
    best_results = calculate_dcf(best_inputs)
    
    # Worst Case: -20% Wachstum, +0.5% WACC
    worst_inputs = inputs.model_copy(deep=True)
    worst_inputs.revenue_growth_rate = inputs.revenue_growth_rate * 0.8
    worst_inputs.terminal_growth = max(inputs.terminal_growth * 0.8, 0.005)  # Min 0.5%
    worst_inputs.risk_free_rate = inputs.risk_free_rate + 0.005
    worst_results = calculate_dcf(worst_inputs)
    
    return {
        "base_case": {
            "name": "Base Case",
            "share_price": base_results["share_price"],
            "enterprise_value": base_results["enterprise_value"],
            "wacc": base_results["wacc"],
            "revenue_growth": inputs.revenue_growth_rate,
            "terminal_growth": inputs.terminal_growth
        },
        "best_case": {
            "name": "Best Case",
            "share_price": best_results["share_price"],
            "enterprise_value": best_results["enterprise_value"],
            "wacc": best_results["wacc"],
            "revenue_growth": best_inputs.revenue_growth_rate,
            "terminal_growth": best_inputs.terminal_growth
        },
        "worst_case": {
            "name": "Worst Case",
            "share_price": worst_results["share_price"],
            "enterprise_value": worst_results["enterprise_value"],
            "wacc": worst_results["wacc"],
            "revenue_growth": worst_inputs.revenue_growth_rate,
            "terminal_growth": worst_inputs.terminal_growth
        }
    }

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
